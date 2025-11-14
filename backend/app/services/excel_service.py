from io import BytesIO
from pathlib import Path
from typing import Iterable, Tuple, Optional, cast
from datetime import datetime

from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Border, Side
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.utils.protection import hash_password
from openpyxl.cell.cell import MergedCell
from openpyxl.styles.protection import Protection as CellProtection

from ..models import OI, Bancada
from ..core.settings import get_settings
from .rules_service import pma_to_pressure, find_exact_in_range, normalize_for_excel_list


HEADER_ROW = 8
DATA_START_ROW = 9
SHEET_NAME = "ERROR FINAL"
Q3_RANGE = "AZ2:BC2"       # lista para E4
ALCANCE_RANGE = "AZ1:BE1"  # lista para O4
FORMULA_START_COL = "AU"
FORMULA_END_COL = "BL"

def _find_header_col(ws: Worksheet, header_name: str, header_row: int = HEADER_ROW) -> Optional[int]:
    """
    Busca la columna de una cabecera por texto exacto (case-insensitive) en la fila `header_row`.
    Retorna el índice de columna (1..N) o None si no existe.
    """
    target = header_name.strip().lower()
    for cell in ws[header_row]:
        text = "" if cell.value is None else str(cell.value).strip().lower()
        if text == target:
            # openpyxl ofrece cell.col_idx (int). SI no, convertimos desde .column (letra o int).
            col = getattr(cell, "col_idx", None)
            if isinstance(col, int):
                return col
            raw = getattr(cell, "column", None)
            if isinstance(raw, int):
                return raw
            if isinstance(raw, str):
                return column_index_from_string(raw)
    return None

def _ensure_workbook() -> Tuple[Workbook, Worksheet]:
    settings = get_settings()
    tpl = Path(settings.template_abs_path)
    if tpl.exists():
        # Mantener vínculos externos tal cual en la plantilla para evitar
        # los avisos de “reparaciones” al abrir en Excel.
        wb = load_workbook(tpl, data_only=False, keep_links=True)
        active = wb.active or (wb.worksheets[0] if wb.worksheets else None)
        if active is None:
            wb.create_sheet("Sheet1")
            active = wb.worksheets[0]
        ws = cast(Worksheet, active)
    else:
        # Fallback para no bloquear pruebas si la plantilla no está
        wb = Workbook()
        active = wb.active or (wb.worksheets[0] if wb.worksheets else None)
        if active is None:
            wb.create_sheet("Sheet1")
            active = wb.worksheets[0]
        ws = cast(Worksheet, active)
        ws["A8"] = "Item"
        ws["B8"] = "# Medidor"
        ws["C8"] = "Estado"
    return wb, ws

def _get_sheet(wb: Workbook, name: str) -> Worksheet:
    if name in wb.sheetnames:
        return wb[name]
    active = wb.active or (wb.worksheets[0] if wb.worksheets else None)
    if active is None:
        wb.create_sheet("Sheet1")
        active = wb.worksheets[0]
    return cast(Worksheet, active)

def _iter_range_values(ws: Worksheet, range_ref: str) -> list[str]:
    """Devuelve los valores del rango ya normalizados al formato de la lista (coma y 1 decimal)."""
    vals: list[str] = []
    for row in ws[range_ref]:
        for cell in row:
            # normalize_for_excel_list puede devolver None → forzamos str
            vals.append(normalize_for_excel_list(cell.value) or "")
    return vals

def _copy_row_styles(src_ws: Worksheet, src_row: int, dst_ws: Worksheet, dst_row: int, max_col: int) -> None:
    """Replica estilos de la fila `src_row` en `dst_row` (A..max_col).
    Evita celdas fusionadas y clona `protection` para no pasar un StyleProxy."""
    # altura de fila idéntica (si aplica)
    try:
        dst_ws.row_dimensions[dst_row].height = src_ws.row_dimensions[src_row].height
    except Exception:
        pass
    for c in range(1, max_col + 1):
        src = src_ws.cell(row=src_row, column=c)
        dst = dst_ws.cell(row=dst_row, column=c)
        # no asignar estilo a celdas fusionadas no-ancla
        if isinstance(dst, MergedCell):
            continue
        if src.has_style:
            try:
                dst.font = src.font.copy()
                dst.number_format = src.number_format
                dst.alignment = src.alignment.copy()
                dst.fill = src.fill.copy()
                dst.border = src.border.copy()
                # Protection: clonar; evitar StyleProxy
                prot = getattr(src, "protection", None)
                if prot is not None:
                    try:
                        dst.protection = prot.copy()
                    except Exception:
                        dst.protection = CellProtection(locked=prot.locked, hidden=prot.hidden)
            except Exception:
                # si algo falla en una celda, continuar con la siguiente
                continue

def _apply_thick_bottom_border(ws: Worksheet, row: int, start_col_letter: str, end_col_letter: str) -> None:
    start_col = column_index_from_string(start_col_letter)
    end_col = column_index_from_string(end_col_letter)
    thick = Side(style="thick")
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        b = cell.border
        cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=thick)

def _copy_formulas(ws: Worksheet, src_row: int, dst_row: int, start_col_letter: str, end_col_letter: str) -> None:
    start_col = column_index_from_string(start_col_letter)
    end_col = column_index_from_string(end_col_letter)
    for c in range(start_col, end_col + 1):
        src_cell = ws.cell(row=src_row, column=c)
        dst_cell = ws.cell(row=dst_row, column=c)
        # Evitar escribir en celdas fusionadas (solo la ancla acepta value)
        if isinstance(dst_cell, MergedCell):
            continue
        if src_cell.data_type == "f" or (isinstance(src_cell.value, str) and str(src_cell.value).startswith("=")):
            formula = str(src_cell.value)
            dst_cell.value = formula.replace(str(src_row), str(dst_row))
        else:
            try:
                dst_cell.value = src_cell.value
            except AttributeError:
                # Si el destino está mergeado/protegido, omitir
                pass 

def generate_excel(oi: OI, bancadas: Iterable[Bancada], password: str | None = None) -> Tuple[bytes, str]:
    wb, _ws_active = _ensure_workbook()
    ws = _get_sheet(wb, SHEET_NAME)  # usar siempre "ERROR FINAL"

    # Celdas fijas de cabecera (selección exacta desde listas)
    q3_candidates = _iter_range_values(ws, Q3_RANGE)
    alcance_candidates = _iter_range_values(ws, ALCANCE_RANGE)
    # normalize_for_excel_list puede devolver None → forzamos str con ""
    q3_value = find_exact_in_range(q3_candidates, normalize_for_excel_list(oi.q3) or "")
    alcance_value = find_exact_in_range(alcance_candidates, normalize_for_excel_list(oi.alcance) or "")
    if q3_value is None:
        raise ValueError("Q3 no coincide con la lista de la plantilla")
    if alcance_value is None:
        raise ValueError("Alcance no coincide con la lista de la plantilla")
    ws["E4"] = q3_value
    ws["O4"] = alcance_value

    # Asegurar cabecera "Estado" (fila 8)
    estado_col = _find_header_col(ws, "Estado", header_row=HEADER_ROW)
    if estado_col is None:
        last_col = ws.max_column + 1
        ws.cell(row=HEADER_ROW, column=last_col, value="Estado")
        estado_col = last_col

    # Otras columnas conocidas (opcionales)
    medidor_col = _find_header_col(ws, "# Medidor", header_row=HEADER_ROW)

    # Escribir filas desde la 9
    rows = list(bancadas)
    # Ordenar por item si existe
    rows.sort(key=lambda b: (b.item or 0))
    
    # Datos globales para columnas B, C, D, E, H
    today_str = datetime.now().strftime("%Y-%m-%d")
    presion_val = pma_to_pressure(oi.pma) if oi.pma else None

    current_row = DATA_START_ROW
    for i, b in enumerate(rows, start=1):
        # 1. Detectar fuente de filas: ¿Tiene data del Grid (rows_data) o es legacy?
        rows_source = getattr(b, "rows_data", []) or []
        nrows = len(rows_source) if rows_source else int(getattr(b, "rows", 15) or 15)

        # 2. Iterar fila por fila
        for k in range(nrows):
            r = current_row + k
            # Obtener payload de la fila k (si existe)
            row_payload = rows_source[k] if (rows_source and k < len(rows_source)) else {}

            # Col A: Item incremental
            item_value = current_row - DATA_START_ROW + 1 + k
            ws.cell(row=r, column=1, value=item_value)

            # Col B y C: Fechas
            ws.cell(row=r, column=2, value=today_str)
            ws.cell(row=r, column=3, value=today_str)

            # Col D y E: Banco y Técnico
            ws.cell(row=r, column=4, value=oi.banco_id)
            ws.cell(row=r, column=5, value=oi.tech_number)

            # Col G: Medidor (Prioridad: Fila > Bancada > Vacío)
            val_medidor = row_payload.get("medidor") or b.medidor
            if medidor_col:
                ws.cell(row=r, column=medidor_col, value=val_medidor or "")

            # Col H: Presión
            if presion_val is not None:
                ws.cell(row=r, column=8, value=presion_val)

            # Col I: Estado
            ws.cell(row=r, column=estado_col, value=(b.estado if b.estado is not None else 0))

            # --- ESCRITURA DE BLOQUES Q3 / Q2 / Q1 ---
            def _write_block(start_col, block):
                if not block: return
                # c1..c7 -> offsets 0..6
                for idx, key in enumerate(["c1", "c2", "c3", "c4", "c5", "c6", "c7"]):
                    val = block.get(key)
                    if val is not None:
                        ws.cell(row=r, column=start_col + idx, value=val)

            _write_block(10, row_payload.get("q3")) # J=10
            _write_block(22, row_payload.get("q2")) # V=22
            _write_block(34, row_payload.get("q1")) # AH=34

            # Copiar/ajustar fórmulas AU:BL desde fila 9 (plantilla base)
            _copy_formulas(ws, DATA_START_ROW, r, FORMULA_START_COL, FORMULA_END_COL)
            # Replicar estilo (bordes, fuente, locked)
            _copy_row_styles(ws, DATA_START_ROW, ws, r, column_index_from_string(FORMULA_END_COL))
        
        # Actualizar puntero global de filas
        current_row += nrows

        # Borde inferior grueso de A a BL en la última fila de la bancada
        _apply_thick_bottom_border(ws, current_row - 1, "A", FORMULA_END_COL)
    
    # Proteger libro/estructura y hojas (usar hash en el workbook)
    if password:
        wb.security = WorkbookProtection(lockStructure=True)
        wb.security.workbookPassword = hash_password(password)  # <— HASH correcto
        for sheet in wb.worksheets:
            sheet.protection.set_password(password)
            sheet.protection.enable()
    
    
    # Guardar en memoria
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{oi.code}.xlsx"
    return buf.read(), filename
